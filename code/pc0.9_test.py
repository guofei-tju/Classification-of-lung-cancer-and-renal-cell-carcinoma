import numpy as np
from sklearn import svm
from sklearn import metrics
from hyperopt import fmin, tpe, hp, STATUS_OK, Trials
from sklearn import model_selection
from openpyxl import load_workbook
import optunity
from imblearn.metrics import specificity_score
from sklearn.ensemble import RandomForestClassifier
from read import read_file_np1
from read import read_label

#导入数据
methods_name = ['gene', 'isoform', 'meth']
disease = ['KIDNEY', 'LUNG']
j = 2
for di in range(2):
    for it in range(3):
        dis = disease[di]
        print(dis + ':')
        name = methods_name[it]
        print(name + ':')
        sample = read_file_np1('E:\multiple_classification\multi_classification\pca\\' + dis + '_' + name + '.csv')
        label = read_label('E:\multiple_classification\data\\'+ dis + '\label\label'+str(di+1)+'.csv')
        #print(file.shape)
        sample = np.array(sample.T)

        if sample.ndim == 1:
            sample = sample.reshape(-1, 1)
        print(sample.shape)
        print(label.shape)

        cv_decorator = optunity.cross_validated(x=sample, y=label, num_folds=5)


        def svr_rforest_tuned_auroc(x_train, y_train, x_test, y_test, n_estimators, max_depth, min_samples_leaf,
                                    min_samples_split):
            rf = RandomForestClassifier(n_estimators=int(n_estimators), max_features='sqrt',
                                        max_depth=int(max_depth), min_samples_leaf=int(min_samples_leaf),
                                        min_samples_split=int(min_samples_split)).fit(x_train, y_train)
            y_pre = rf.predict(x_test)
            #pcc = round(np.corrcoef(y_pre, y_test)[0][1], 5)
            accuracy = metrics.accuracy_score(y_test, y_pre)
            return accuracy
            # auc = optunity.metrics.roc_auc(y_test, decision_values)
            # print(pcc_test)
            # return optunity.metrics.mse(y_test, y_pre)


        svr_rforest_tuned_auroc = cv_decorator(svr_rforest_tuned_auroc)
        # this is equivalent to the more common syntax below
        # @optunity.cross_validated(x=data, y=labels, num_folds=5)
        # def svm_rbf_tuned_auroc...max_features=['square', 'log'],

        optimal_rbf_pars, info, _ = optunity.maximize(svr_rforest_tuned_auroc, num_evals=200, n_estimators=[1, 200],
                                                      max_depth=[1, 100], min_samples_leaf=[1, 20],
                                                      min_samples_split=[2, 20])
        # when running this outside of IPython we can parallelize via optunity.pmap
        # optimal_rbf_pars, _, _ = optunity.maximize(svm_rbf_tuned_auroc, 150, C=[0, 10], gamma=[0, 0.1], pmap=optunity.pmap)

        print("Optimal parameters: " + str(optimal_rbf_pars))
        print("PCC of tuned rf with hyper parameters %1.5f" % info.optimum)
        rf = RandomForestClassifier(n_estimators=int(optimal_rbf_pars['n_estimators']), max_features='sqrt',
                                     max_depth=int(optimal_rbf_pars['max_depth']), min_samples_leaf=
                                     int(optimal_rbf_pars['min_samples_leaf']), min_samples_split=
                                     int(optimal_rbf_pars['min_samples_split']))

        cv = model_selection.StratifiedKFold(n_splits=5, shuffle=True, random_state=0)

        if dis == 'LUNG':
            # lung五折交叉验证
            scorerMCC = metrics.make_scorer(metrics.matthews_corrcoef)
            scorerSP = metrics.make_scorer(specificity_score)
            scorerPR = metrics.make_scorer(metrics.precision_score)
            scorerSE = metrics.make_scorer(metrics.recall_score)

            scorer = {'ACC': 'accuracy', 'recall': scorerSE, 'roc_auc': 'roc_auc', 'MCC': scorerMCC, 'SP': scorerSP}
            five_fold = model_selection.cross_validate(rf, sample, label, cv=cv, scoring=scorer)

            mean_ACC = np.mean(five_fold['test_ACC'])
            mean_sensitivity = np.mean(five_fold['test_recall'])
            mean_AUC = np.mean(five_fold['test_roc_auc'])
            mean_MCC = np.mean(five_fold['test_MCC'])
            mean_SP = np.mean(five_fold['test_SP'])

            wb = load_workbook("rf_0.8.xlsx")  # 生成一个已存在的wookbook对象
            wb1 = wb.active  # 激活sheet
            wb1.cell(j, 2, round(mean_sensitivity, 5))
            wb1.cell(j, 3, round(mean_SP, 5))
            wb1.cell(j, 4, round(mean_ACC, 5))
            wb1.cell(j, 5, round(mean_MCC, 5))
            wb1.cell(j, 6, round(mean_AUC, 5))
            wb1.cell(j, 7, dis)
            wb1.cell(j, 8, name)
            wb.save("rf_0.8.xlsx")  # 保存

            print('five fold:')
            print('SN =', round(mean_sensitivity, 5))
            print('SP =', round(mean_SP, 5))
            print('ACC =', round(mean_ACC, 5))
            print('MCC = ', round(mean_MCC, 5))
            print('AUC = ', round(mean_AUC, 5))
            j = j + 1
        else:
            # kidney五折交叉验证
            scorerMCC = metrics.make_scorer(metrics.matthews_corrcoef)
            scorerSP = metrics.make_scorer(specificity_score, average='micro')
            # scorerPR = metrics.make_scorer(metrics.precision_score)
            scorerSE = metrics.make_scorer(metrics.recall_score, average='micro')

            scorer = {'ACC': 'accuracy', 'recall': scorerSE, 'roc_auc': 'roc_auc_ovo', 'SP': scorerSP, 'MCC': scorerMCC}
            five_fold = model_selection.cross_validate(rf, sample, label, cv=cv, scoring=scorer)
            print(five_fold)
            mean_ACC = np.mean(five_fold['test_ACC'])
            mean_sensitivity = np.mean(five_fold['test_recall'])
            mean_AUC = np.mean(five_fold['test_roc_auc'])
            mean_MCC = np.mean(five_fold['test_MCC'])
            mean_SP = np.mean(five_fold['test_SP'])

            wb = load_workbook("rf_0.8.xlsx")  # 生成一个已存在的wookbook对象
            wb1 = wb.active  # 激活sheet
            wb1.cell(j, 2, round(mean_sensitivity, 5))
            wb1.cell(j, 3, round(mean_SP, 5))
            wb1.cell(j, 4, round(mean_ACC, 5))
            wb1.cell(j, 5, round(mean_MCC, 5))
            wb1.cell(j, 6, round(mean_AUC, 5))
            wb1.cell(j, 7, dis)
            wb1.cell(j, 8, name)
            wb.save("rf_0.8.xlsx")  # 保存

            print('five fold:')
            print('SN =', round(mean_sensitivity, 5))
            print('SP =', round(mean_SP, 5))
            print('ACC =', round(mean_ACC, 5))
            print('MCC = ', round(mean_MCC, 5))
            print('AUC = ', round(mean_AUC, 5))
            j = j + 1